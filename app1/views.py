from django.shortcuts import HttpResponse, render, redirect
from wordcloud import WordCloud
from app1 import models
from django.utils.safestring import mark_safe
from openpyxl import load_workbook
from django import forms
from django.db.models import Count, Q
import json
from django.db.models.functions import Cast, Substr
from django.db.models import DateField
import os
from mywork_1.settings import BASE_DIR
from django.utils import timezone
import jieba.analyse
from snownlp import SnowNLP
import jieba
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
import nltk
from nltk.corpus import stopwords
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.metrics import confusion_matrix, roc_curve, auc, average_precision_score, ndcg_score, classification_report
import matplotlib.pyplot as plt
import numpy as np
from sklearn.preprocessing import label_binarize
from datetime import datetime


def register(request):
    # 跳转到注册界面
    if request.method == "GET":
        return render(request, 'register.html')
    # 获取输入信息
    name = request.POST.get("name")
    password_ = request.POST.get("password")
    # 规范输入
    if name == "" or password_ == "":
        return HttpResponse("请按照要求输入")
    # 检测用户是否已经注册过
    obj = models.user.objects.filter(name=name).first()
    # 新用户将数据存储到数据库，root=2表示普通用户
    if obj is None:
        models.user.objects.create(name=name, password=password_, root=2)
        return redirect('/login/')
    # 已经注册过的用户不允许重复注册
    else:
        return HttpResponse("用户已存在！")


# 登录
def login(request):
    # 跳转到登陆页面
    if request.method == "GET":
        return render(request, 'login.html')
    # 获取登录信息
    name = request.POST.get("name")
    password_ = request.POST.get("password")
    # 检测是否是注册过的用户
    obj = models.user.objects.filter(name=name).first()
    # 该用户没有注册过
    if obj is None:
        return render(request, 'login.html', {"field": "不存在用户"})
    # 密码正确且是管理员
    if obj.password == password_ and obj.root == 1:
        return redirect("/admin_show/")
    # 密码正确的普通用户
    elif obj.password == password_ and obj.root == 2:
        return redirect("/user_show/")
    # 密码错误
    else:
        return render(request, 'login.html', {"field": "密码错误"})


# 管理员登录后展示界面
def admin_show(request):
    # 根据请求参数中的“q”参数过滤信息列表
    data_dict = {}
    value = request.GET.get('q')
    value_c = request.GET.get('c')
    # 如果“q”参数与”c“参数存在，就将一个key-value对添加到data_dict中
    # key为“title__contains”，value为q或c的值
    if value:
        data_dict["Title__contains"] = value
    if value_c:
        # print(value_c)
        data_dict["City__contains"] = value_c
    # 分页
    # 获取请求参数中的“page”参数
    # page变量表示当前页码，page_size变量表示每页信息的数量
    page = int(request.GET.get('page', 1))
    page_size = 13
    start = (page - 1) * page_size
    end = page * page_size
    # filter()方法会根据data_dict中的条件来过滤信息列表
    # “title__contains”是一个特殊的字段查询方式，表示查询title字段中包含某个字符串的信息记录
    total_count = models.information.objects.filter(**data_dict).order_by("-id").count()
    queryset = models.information.objects.filter(**data_dict).order_by("-id")[start:end]
    # 计算出总页数total_page_count，以及显示在页面上的页码范围
    # 其中，plus变量表示当前页码前后显示的页码数量
    total_page_count, div = divmod(total_count, page_size)
    if div:
        total_page_count += 1
    plus = 5
    if total_page_count <= 2 * plus + 1:
        start_page = 1
        end_page = total_page_count + 1
    else:
        if page <= plus:
            start_page = 1
            end_page = 2 * plus + 1
        else:
            if (page + plus) > total_page_count:
                start_page = total_page_count - 2 * plus
                end_page = total_page_count + 1
            else:
                start_page = page - plus
                end_page = page + plus + 1
    # 使用了字符串的format()方法来生成HTML代码，其中{}表示占位符，用于将后面的参数插入到字符串中
    page__str_list = []
    prev = '<li><a href="?page={}">首页</a></li>'.format(1)
    page__str_list.append(prev)
    # 首先根据当前页码page判断是否需要生成“上一页”链接
    if page > 1:
        prev = '<li><a href="?page={}">«</a></li>'.format(page - 1)
    else:
        prev = '<li><a href="?page={}">«</a></li>'.format(1)
    # 将生成的HTML代码用join()方法连接起来，生成完整的分页导航栏HTML代码
    page__str_list.append(prev)
    for i in range(start_page, end_page):
        # 如果当前页码等于循环中的页码i，则生成带有“active”类的HTML代码，表示当前页码。否则，生成普通的HTML代码
        if i == page:
            ele = '<li class="active"><a href="?page={}">{}</a></li>'.format(i, i)
        else:
            ele = '<li><a href="?page={}">{}</a></li>'.format(i, i)
        page__str_list.append(ele)
    # 根据当前页码page和总页数total_page_count判断是否需要生成“下一页”链接
    if page < total_page_count:
        prev = '<li><a href="?page={}">»</a></li>'.format(page + 1)
    else:
        prev = '<li><a href="?page={}">»</a></li>'.format(total_page_count)
    page__str_list.append(prev)
    # 生成分页导航栏中的“末页”链接的HTML代码
    prev = '<li><a href="?page={}">末页</a></li>'.format(total_page_count)
    page__str_list.append(prev)
    # 这段代码将生成的分页导航栏HTML代码转换为Django模板中的safe字符串，以避免在模板渲染时被转义
    # Django的mark_safe()方法将page__str_list中的HTML代码连接起来，并将结果转换为safe字符串
    page_string = mark_safe("".join(page__str_list))
    # 将分页信息渲染到infro__show.html中
    return render(request, 'admin_show.html', {'queryset': queryset, 'page_string': page_string})


# 普通用户登陆后的界面
def user_show(request):
    data_dict = {}
    value = request.GET.get('q')
    value_c = request.GET.get('c')
    if value_c:
        # print(value_c)
        data_dict["City__contains"] = value_c
    if value:
        data_dict["Title__contains"] = value
    # 分页
    page = int(request.GET.get('page', 1))
    page_size = 13
    start = (page - 1) * page_size
    end = page * page_size
    total_count = models.information.objects.filter(**data_dict).order_by("-id").count()
    queryset = models.information.objects.filter(**data_dict).order_by("-id")[start:end]
    total_page_count, div = divmod(total_count, page_size)
    if div:
        total_page_count += 1
    plus = 5
    if total_page_count <= 2 * plus + 1:
        start_page = 1
        end_page = total_page_count + 1
    else:
        if page <= plus:
            start_page = 1
            end_page = 2 * plus + 1
        else:
            if (page + plus) > total_page_count:
                start_page = total_page_count - 2 * plus
                end_page = total_page_count + 1
            else:
                start_page = page - plus
                end_page = page + plus + 1
    page__str_list = []
    prev = '<li><a href="?page={}">首页</a></li>'.format(1)
    page__str_list.append(prev)
    if page > 1:
        prev = '<li><a href="?page={}">«</a></li>'.format(page - 1)
    else:
        prev = '<li><a href="?page={}">«</a></li>'.format(1)
    page__str_list.append(prev)
    for i in range(start_page, end_page):
        if i == page:
            ele = '<li class="active"><a href="?page={}">{}</a></li>'.format(i, i)
        else:
            ele = '<li><a href="?page={}">{}</a></li>'.format(i, i)
        page__str_list.append(ele)
    if page < total_page_count:
        prev = '<li><a href="?page={}">»</a></li>'.format(page + 1)
    else:
        prev = '<li><a href="?page={}">»</a></li>'.format(total_page_count)
    page__str_list.append(prev)
    prev = '<li><a href="?page={}">末页</a></li>'.format(total_page_count)
    page__str_list.append(prev)
    page_string = mark_safe("".join(page__str_list))
    return render(request, 'user_show.html', {'queryset': queryset, 'page_string': page_string})


# 管理员添加一条数据
def admin_add(request):
    if request.method == "GET":
        return render(request, 'admin_add.html')
    QueSub = request.POST.get("QueSub")
    City = request.POST.get("City")
    Title = request.POST.get("Title")
    Class = request.POST.get("Class")
    Time = request.POST.get("Time")
    RepTime = request.POST.get("RepTime")
    Content = request.POST.get("Content")
    RepDep = request.POST.get("RepDep")
    RepCon = request.POST.get("RepCon")
    URL = request.POST.get("URL")
    reading = 0
    # 保存到数据库
    models.information.objects.create(
        QueSub=QueSub,
        City=City,
        Title=Title,
        Class=Class,
        Time=Time,
        RepTime=RepTime,
        Content=Content,
        RepDep=RepDep,
        RepCon=RepCon,
        URL=URL,
        reading=reading,
    )
    return redirect("/admin_show/")


# 管理员批量上传数据
def admin_upload(request):
    if request.method == "GET":
        return render(request, "admin_upload.html")
    file_obj = request.FILES.get("upload")
    return HttpResponse("上传成功")


# 后台处理批量上传的数据
def admin_multi(request):
    # 批量上传
    if request.method == 'POST':
        file_obj = request.FILES.get("upload")
        if not file_obj:
            return render(request, 'admin_upload_error.html', {'error': '您未选择文件！'})
        try:
            # 使用 openpyxl 库的 load_workbook 方法加载 Excel 文件
            wb = load_workbook(file_obj)
        except Exception as e:
            return render(request, 'admin_upload_error.html', {'error': '您上传的文件格式不正确!'})
        # 获取 Excel 文件的第一个工作表
        sheet = wb.worksheets[0]
        # 统计数据上传情况
        success_count = 0
        fail_count = 0
        # 使用 iter_rows 方法遍历每一行数据，从第二行开始遍历（因为第一行通常是表头）
        for row in sheet.iter_rows(min_row=2):
            # 检测一下
            # for i in range(0,10):
            # print(row[i].value)
            exist = models.information.objects.filter(
                QueSub=row[0].value,
                City=row[1].value,
                Title=row[2].value,
                Class=row[3].value,
                Time=row[4].value,
                Content=row[5].value,
                RepTime=row[7].value,
                RepDep=row[6].value,
                RepCon=row[8].value,
                URL=row[9].value,
                reading=0,
            ).exists()
            if exist:
                fail_count += 1
            if not exist:
                try:
                    models.information.objects.create(
                        QueSub=row[0].value,
                        City=row[1].value,
                        Title=row[2].value,
                        Class=row[3].value,
                        Time=row[4].value,
                        Content=row[5].value,
                        RepTime=row[7].value,
                        RepDep=row[6].value,
                        RepCon=row[8].value,
                        URL=row[9].value,
                        reading=0,
                    )
                    success_count += 1
                except Exception as e:
                    # print(f"Error: {str(e)}. Skip this row.")
                    fail_count += 1
                    continue
        return render(request, 'admin_upload_error.html', {'success_count': success_count, 'fail_count': fail_count})


# 管理员对某一条数据进行编辑
def admin_edit(request, id):
    if request.method == "GET":
        # 测试
        # print(id)
        # 获取指定数据元组信息
        obj = models.information.objects.filter(id=id).first()
        return render(request, 'admin_edit.html', {"obj": obj})
    # 修改数据库中指定数据信息
    QueSub = request.POST.get("QueSub")
    City = request.POST.get("City")
    Title = request.POST.get("Title")
    Class = request.POST.get("Class")
    Time = request.POST.get("Time")
    RepTime = request.POST.get("RepTime")
    Content = request.POST.get("Content")
    RepDep = request.POST.get("RepDep")
    RepCon = request.POST.get("RepCon")
    URL = request.POST.get("URL")
    reading = request.POST.get("reading")
    models.information.objects.filter(id=id).update(
        QueSub=QueSub,
        City=City,
        Title=Title,
        Class=Class,
        Time=Time,
        RepTime=RepTime,
        Content=Content,
        RepDep=RepDep,
        RepCon=RepCon,
        URL=URL,
        reading=reading,
    )
    return redirect("/admin_show/")


# 管理员删除某一条数据
def admin_delete(request, id):
    models.information.objects.filter(id=id).delete()
    return redirect("/admin_show/")


# 查看信息可视化大屏
def user_display(request):
    queryset1 = models.information.objects.all()
    # 处理全局扇形图global_chart_1数据
    data1 = global_chart1_data(queryset1)
    # print(data1)
    # 全局扇形图global_chart_2数据
    data2 = global_chart2_data(queryset1)
    # print(data2)
    # 处理最新数据展示
    data_lasted = queryset1.order_by('-Time')[:6]
    data_count = queryset1.count()
    # 全局情感分析折线图global_chart_3
    data3 = global_chart3_data(queryset1)
    # print(data3)
    # 全局词云图global_chart_4\global_chart_5\global_chart_6，生成词云图并存储到static/img中
    global_chart4_data(queryset1)
    # 将数据传回user_charts页面
    return render(
        request, 'user_display.html', {
            'data1': data1,
            'data2': data2,
            'data_lasted': data_lasted,
            'data_count': data_count,
            'data3': data3,
        }
    )


# 全局扇形图global_chart_1数据
def global_chart1_data(queryset1):
    data1_temp = queryset1.values("City").annotate(count1=Count("id")).all()
    # print(data1_temp)
    data1_json = json.dumps(list(data1_temp), default=str, ensure_ascii=False)
    # print(data1_json)
    return data1_json


# 全局扇形图global_chart_2数据
def global_chart2_data(queryset1):
    # 只截取年份
    data2_temp = queryset1.annotate(date=Substr("Time", 1, 4)).values("date").annotate(count2=Count("id")).order_by(
        "date").all()
    # print(data2_temp)
    data2_json = json.dumps(list(data2_temp), default=str, ensure_ascii=False)
    # print(data2_json)
    return data2_json


# 全局词云图global_chart_4，生成词云图并存储到static/img中
def global_chart4_data(queryset1):
    # 只截取年份与月份
    # .exclude(xxx__isnull=True)用于排除字段为空的情况
    # 五十万条数据全跑时间太长，只取最近的一万条
    queryset1 = queryset1.exclude(Title__isnull=True).exclude(Content__isnull=True).exclude(
        RepCon__isnull=True).order_by('-Time').all()[: 10000]
    # 获取标题Title数组和内容content数组
    global_titles_temp1 = [item.Title for item in queryset1]
    global_contents_temp1 = [item.Content for item in queryset1]
    global_repcon_temp1 = [item.RepCon for item in queryset1]
    # 分别合并为一整个字符串
    global_titles_combined1 = ''.join(global_titles_temp1)
    global_contents_combined1 = ''.join(global_contents_temp1)
    global_repcons_combined1 = ''.join(global_repcon_temp1)
    # jieba分词
    global_titles = jieba.cut(global_titles_combined1)
    global_contents = jieba.cut(global_contents_combined1)
    global_repcons = jieba.cut(global_repcons_combined1)
    # 定义要忽略的单词列表
    stopwords = set()
    # 获取当前脚本的绝对路径
    # script_dir = os.path.dirname(os.path.abspath(__file__))
    # with open(os.path.join(script_dir, 'static/txt/stopwords.txt'), 'r', encoding='utf-8') as f:
    # 读取第一个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/baidu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第二个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/cn_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第三个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/hit_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第四个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/scu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 将自己的也加入
    with open('static/txt/my_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 输出所有的停用词
    # print(stopwords)
    # 使用 WordCloud 库生成标题Title词云图
    global_chart4_title = WordCloud(
        width=800,
        height=400,
        background_color='white',
        max_words=40,
        stopwords=stopwords,
    ).generate(' '.join(global_titles))
    # 使用 WordCloud 库生成Content词云图
    global_chart4_content = WordCloud(
        width=800,
        height=400,
        background_color='white',
        max_words=40,
        stopwords=stopwords,
    ).generate(' '.join(global_contents))
    # 使用 WordCloud 库生成RepCon词云图
    global_chart4_repcon = WordCloud(
        width=800,
        height=400,
        background_color='white',
        max_words=40,
        stopwords=stopwords,
    ).generate(' '.join(global_repcons))
    # 将词云图保存为图片文件
    img_folder = os.path.join(BASE_DIR, 'static', 'img')
    if not os.path.exists(img_folder):
        os.makedirs(img_folder)
    img_repcon_path = os.path.join(img_folder, 'global_chart4_repcon.png')
    global_chart4_repcon.to_file(img_repcon_path)
    img_title_path = os.path.join(img_folder, 'global_chart4_title.png')
    global_chart4_title.to_file(img_title_path)
    img_content_path = os.path.join(img_folder, 'global_chart4_content.png')
    global_chart4_content.to_file(img_content_path)


# 全局情感分析折线图global_chart_3
def global_chart3_data(queryset1):
    global_data3_temp = queryset1.annotate(date3=Substr("Time", 1, 7)).values("date3", "Content").order_by(
        "-date3").all()[:200000]
    # 导入中文情感词典
    global_positive_dict = set()
    global_negative_dict = set()
    with open('static/txt2/qinghua_positive.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_positive_dict.update(word)
    with open('static/txt2/taiwan_positive.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_positive_dict.update(word)
    with open('static/txt2/zhiwang_positive.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_positive_dict.update(word)
    with open('static/txt2/zhiwang_positive2.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_positive_dict.update(word)
    with open('static/txt2/qinghua_negative.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_negative_dict.update(word)
    with open('static/txt2/taiwan_negative.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_negative_dict.update(word)
    with open('static/txt2/zhiwang_negative.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_negative_dict.update(word)
    with open('static/txt2/zhiwang_negative2.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        global_negative_dict.update(word)
    # print(global_positive_dict)
    # print(global_negative_dict)
    # 计算情感分数
    # 数据预处理，依然是jieba分词
    global_stopwords = set()
    with open('static/txt/baidu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        global_stopwords.update(words)
    # 读取第二个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/cn_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        global_stopwords.update(words)
    # 读取第三个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/hit_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        global_stopwords.update(words)
    # 读取第四个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/scu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        global_stopwords.update(words)
    # 将自己的也加入
    with open('static/txt/my_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        global_stopwords.update(words)
    global_stopwords_file = 'static/txt2/global_my_stopwords.txt'
    # 将停用词写入文件
    with open(global_stopwords_file, 'w', encoding='utf-8') as f:
        for word in global_stopwords:
            f.write(word + '\n')
    # 设置jieba的停用词
    jieba.analyse.set_stop_words(global_stopwords_file)
    print(global_data3_temp)
    global_data3_sentiments = []  # 存储日期和情感分数的列表
    for con in global_data3_temp:
        content = con['Content']
        date3 = con['date3']
        # 对内容进行分词
        if content:
            words = jieba.lcut(content)
        # 对每个分词进行情感分析，并计算平均分数
        sentiments = [SnowNLP(word).sentiments for word in words]
        avg_sentiment = sum(sentiments) / len(sentiments)
        # 将分数和日期存储到字典中，并添加到列表中
        global_data3_sentiments.append({'date3': date3, 'sentiment': avg_sentiment})
    # date3相同的结果聚合
    df = pd.DataFrame(global_data3_sentiments)
    result = df.groupby('date3')['sentiment'].mean().reset_index()
    # print(result, type(result))
    # 将DataFrame对象转换为JSON格式的字符串
    data3_json = result.to_json(orient='records')
    print(data3_json)
    return data3_json


# 查看信息分析图表
def user_charts(request):
    return render(request, 'user_charts.html')


# user_charts.html页面各种图表生成位置，需要调用下面四个函数
from datetime import datetime
from django.utils import timezone


def api_data(request):
    if request.method == 'GET':
        city = request.GET.get('city')
        begin_date = request.GET.get('begin_date')
        end_date = request.GET.get('end_date')
        if not begin_date:
            begin_date = '2007-11-02'
        if not end_date:
            end_date = '2022-08-24'
        begin = timezone.make_aware(datetime.datetime.strptime(begin_date, "%Y-%m-%d"), timezone.get_default_timezone())
        end = timezone.make_aware(datetime.datetime.strptime(end_date, "%Y-%m-%d"), timezone.get_default_timezone())
        print(city, begin, end)
        if city == "all":
            # queryset = models.information.objects.all()
            queryset = models.information.objects.filter(Q(Time__range=(begin, end))).all()
            # print(queryset)
            # 处理折线图数据
            data1 = chart1_data(queryset)
            # print(data1)
            # 处理柱状图数据
            data2 = chart2_data(queryset)
            # print(data2)
            # 处理饼图数据
            data3 = chart3_data(queryset)
            # print(data3)
            # 根据查询结果生成词云图并存储到static/img中
            chart4_data(queryset)
            # 处理情感分析图数据
            data5 = chart5_data(queryset)
            # print(data5)
            # 将数据传回user_charts页面
            return render(
                request, 'user_charts.html', {
                    'data1': data1,
                    'data2': data2,
                    'data3': data3,
                    'data5': data5,
                }
            )
        else:
            data_dict = {}
            data_dict["City__contains"] = city
            # queryset = models.information.objects.filter(**data_dict).all()
            queryset = models.information.objects.filter(**data_dict).filter(Q(Time__range=(begin, end))).all()
            # 处理折线图数据
            data1 = chart1_data(queryset)
            # print(data1)
            # 处理柱状图数据
            data2 = chart2_data(queryset)
            # print(data2)
            # 处理饼图数据
            data3 = chart3_data(queryset)
            # print(data3)
            # 根据查询结果生成词云图并存储到static/img中
            chart4_data(queryset)
            # 处理情感分析图数据
            data5 = chart5_data(queryset)
            # print(data5)
            # 将数据传回user_charts页面
            return render(
                request, 'user_charts.html', {
                    'data1': data1,
                    'data2': data2,
                    'data3': data3,
                    'data5': data5,
                }
            )


# 处理折线图数据（测试时先放200条处理，不然速度很慢）
def chart1_data(queryset):
    data1_temp = queryset.annotate(date1=Cast("Time", output_field=DateField())).values("date1").annotate(
        count1=Count("id")).order_by("date1").all()
    for item in data1_temp:
        if item["date1"]:
            item["date1"] = item["date1"].strftime("%Y-%m-%d")
    # 将data1_temp转换为json格式，否则Time格式不是字符串script无法处理
    data1_json = json.dumps(list(data1_temp), default=str)
    return data1_json


# 处理柱状图数据（测试时先放200条处理，不然速度很慢）
def chart2_data(queryset):
    # 按城市检索所有元组，信息最多，旨于绘制时提问时间与回复时间的间间隔柱状图
    data2_temp = queryset.order_by('Time').all()
    # print(data2_temp)
    # 对每一条数据的时间进行格式化（yyyy-mm-dd）
    for item in data2_temp:
        if item.Time:
            item.Time = item.Time.strftime("%Y-%m-%d")
            # print(item.Time)
        if item.RepTime:
            item.RepTime = item.RepTime.strftime("%Y-%m-%d")
            # print(item.RepTime)
    # 测试
    # print(data2_temp[0].Time)
    # print(data2_temp[0].RepTime)
    # print(type(data2_temp[0].Time))
    # print(type(data2_temp[0].RepTime))
    # 计算每一天的平均处理时间并构建列表
    data2_dict_temp = {}
    for item in data2_temp:
        if item.RepTime:
            # 计算每一条元组的时间差
            date_obj1 = datetime.datetime.strptime(item.Time, "%Y-%m-%d")
            date_obj2 = datetime.datetime.strptime(item.RepTime, "%Y-%m-%d")
            delta = date_obj2 - date_obj1
            days = delta.days
            # print(days)
            # 计算每一天的平均时间差
            if item.Time in data2_dict_temp:
                data2_dict_temp[item.Time]["count"] += 1
                data2_dict_temp[item.Time]["days"] += days
            else:
                data2_dict_temp[item.Time] = {"count": 1, "days": days}
    # print(data2_dict_temp)
    data2_list = []
    for date, value in data2_dict_temp.items():
        count = value["count"]
        days = value["days"]
        if days != 0:
            # 保留两位小数
            count2 = round(count / days, 2)
        else:
            count2 = 0
        data2_list.append({"Time": date, "count2": count2})
    # 将结果转换为 JSON 格式并返回
    data2_json = json.dumps(data2_list, default=str)
    # print(data2_json)
    return data2_json


# 处理饼图数据（测试时先放200条处理，不然速度很慢）
def chart3_data(queryset):
    # 只截取年份与月份
    data3_temp = queryset.annotate(date3=Substr("Time", 1, 7)).values("date3").annotate(count3=Count("id")).order_by(
        "date3").all()
    # print(data3_temp)
    data3_json = json.dumps(list(data3_temp), default=str)
    # print(data3_json)
    return data3_json


# 根据查询结果生成词云图并存储到static/img中
def chart4_data(queryset):
    # 只截取年份与月份
    # .exclude(xxx__isnull=True)用于排除字段为空的情况
    # 五十万条数据全跑时间太长，只取最近的一万条
    queryset = queryset.exclude(Title__isnull=True).exclude(Content__isnull=True).order_by('-Time').all()[: 10000]
    # 获取标题Title数组和内容content数组
    titles_temp = [item.Title for item in queryset]
    contents_temp = [item.Content for item in queryset]
    # 分别合并为一整个字符串
    titles_combined = ''.join(titles_temp)
    contents_combined = ''.join(contents_temp)
    # jieba分词
    titles = jieba.cut(titles_combined)
    contents = jieba.cut(contents_combined)
    # 定义要忽略的单词列表
    stopwords = set()
    # 获取当前脚本的绝对路径
    # script_dir = os.path.dirname(os.path.abspath(__file__))
    # with open(os.path.join(script_dir, 'static/txt/stopwords.txt'), 'r', encoding='utf-8') as f:
    # 读取第一个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/baidu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第二个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/cn_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第三个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/hit_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第四个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/scu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 将自己的也加入
    with open('static/txt/my_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 输出所有的停用词
    # print(stopwords)
    # 使用 WordCloud 库生成标题Title词云图
    chart4_title = WordCloud(
        width=800,
        height=400,
        background_color='white',
        max_words=40,
        stopwords=stopwords,
    ).generate(' '.join(titles))
    # 使用 WordCloud 库生成Content词云图
    chart4_content = WordCloud(
        width=800,
        height=400,
        background_color='white',
        max_words=40,
        stopwords=stopwords,
    ).generate(' '.join(contents))
    # 将词云图保存为图片文件
    img_folder = os.path.join(BASE_DIR, 'static', 'img')
    if not os.path.exists(img_folder):
        os.makedirs(img_folder)
    img_title_path = os.path.join(img_folder, 'chart4_title.png')
    chart4_title.to_file(img_title_path)
    img_content_path = os.path.join(img_folder, 'chart4_content.png')
    chart4_content.to_file(img_content_path)


# 局部情感分析图
'''
SnowNLP返回的情感分数是在0到1之间的浮点数，表示文本情感的正向程度，
数值越大表示情感越正向，越小表示情感越负向。
通常，可以将情感分数分为以下几个区间，并对应相应的情感：
0.9 ~ 1.0 ：非常积极，表示非常开心、满意等极其积极的情感。
0.7 ~ 0.9 ：积极，表示开心、满意等积极情感。
0.5 ~ 0.7 ：中性，表示较为冷静、客观的情感。
0.3 ~ 0.5 ：消极，表示不满、不开心等消极情感。
0.0 ~ 0.3 ：非常消极，表示非常不满、非常不开心等极其消极的情感。
需要注意的是，这只是一种常见的情感分数区间划分方式，
具体的情感分数区间和对应的情感可能会因为不同的数据集、任务等而有所不同。
'''


def chart5_data(queryset):
    data5_temp = queryset.annotate(date5=Substr("Time", 1, 7)).values("date5", "Content").order_by("-date5").all()[
                 : 10000]
    # 导入中文情感词典
    positive_dict = set()
    negative_dict = set()
    with open('static/txt2/qinghua_positive.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        positive_dict.update(word)
    with open('static/txt2/taiwan_positive.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        positive_dict.update(word)
    with open('static/txt2/zhiwang_positive.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        positive_dict.update(word)
    with open('static/txt2/zhiwang_positive2.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        positive_dict.update(word)
    with open('static/txt2/qinghua_negative.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        negative_dict.update(word)
    with open('static/txt2/taiwan_negative.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        negative_dict.update(word)
    with open('static/txt2/zhiwang_negative.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        negative_dict.update(word)
    with open('static/txt2/zhiwang_negative2.txt', 'r', encoding='utf-8', errors='ignore') as f:
        word = f.read().split()
        negative_dict.update(word)
    # print(positive_dict)
    # print(negative_dict)
    # 计算情感分数
    # 数据预处理，依然是jieba分词
    stopwords = set()
    with open('static/txt/baidu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第二个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/cn_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第三个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/hit_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 读取第四个停用词文件，并将其中的单词添加到停用词集合中
    with open('static/txt/scu_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    # 将自己的也加入
    with open('static/txt/my_stopwords.txt', 'r', encoding='utf-8') as f:
        words = f.read().split()
        stopwords.update(words)
    stopwords_file = 'static/txt2/my_stopwords.txt'
    # 将停用词写入文件
    with open(stopwords_file, 'w', encoding='utf-8') as f:
        for word in stopwords:
            f.write(word + '\n')
    # 设置jieba的停用词
    jieba.analyse.set_stop_words(stopwords_file)
    # print(data5_temp)
    data5_sentiments = []  # 存储日期和情感分数的列表
    for con in data5_temp:
        content = con['Content']
        date5 = con['date5']
        # 对内容进行分词
        if content:
            words = jieba.lcut(content)
        # 对每个分词进行情感分析，并计算平均分数
        sentiments = [SnowNLP(word).sentiments for word in words]
        avg_sentiment = sum(sentiments) / len(sentiments)
        # 将分数和日期存储到字典中，并添加到列表中
        data5_sentiments.append({'date5': date5, 'sentiment': avg_sentiment})
    # date5相同的结果聚合
    df = pd.DataFrame(data5_sentiments)
    result = df.groupby('date5')['sentiment'].mean().reset_index()
    # print(result, type(result))
    # 将DataFrame对象转换为JSON格式的字符串
    data5_json = result.to_json(orient='records')
    # print(data5_json)
    return data5_json


# 管理员与用户信息展示入口
def admin_list(request):
    queryset = models.user.objects.all()
    return render(request, 'admin_list.html', {"queryset": queryset})


# 下面是身份信息管理时要用到的公共类
'''
这是一个 Django 中的表单类 userModelForm，用于创建和编辑 user 模型的实例。
该表单继承自 Django 自带的 ModelForm 类，提供了一些默认的表单字段和验证逻辑，
使得创建和编辑模型实例的流程更加简单。
在 Meta 类中定义了该表单对应的模型类 user 和需要显示的字段，
即 name、root 和 password。这些字段会自动对应到表单中的对应输入框中，
并在提交表单时将输入的值保存到相应的模型实例中。
在 __init__ 方法中，通过调用父类的 __init__ 方法，
初始化了表单实例，并对每个表单字段的 widget 属性设置了一个 class 属性，
将其样式设置为 "form-control"，这是 Bootstrap 框架中的一个样式类，
使得表单输入框可以使用 Bootstrap 的样式。
'''


class userModelForm(forms.ModelForm):
    class Meta:
        model = models.user
        fields = ["name", "root", "password"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control"}


# 管理员添加新的用户或管理员
def admin_add_user(request):
    if request.method == "GET":
        form = userModelForm()
        return render(request, 'admin_add_user.html', {"form": form})
    form = userModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/admin_list/')
    return render(request, 'admin_add_user.html', {"form": form})


# 管理员编辑用户或管理员信息
def admin_edit_user(request, id):
    row_obj = models.user.objects.filter(id=id).first()
    if request.method == "GET":
        form = userModelForm(instance=row_obj)
        return render(request, 'admin_edit_user.html', {"form": form})
    form = userModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect('/admin_list/')
    return render(request, 'admin_edit_user.html', {"form": form})


# 管理员删除用户或管理员信息
def admin_delete_user(request, id):
    models.user.objects.filter(id=id).delete()
    return redirect("/admin_list/")


# 信息可视化大屏跳转到“分析结果”
def display_1(request):
    return render(request, 'display_1.html')


# 信息可视化大屏跳转到“文本分类”
def display_2(request):
    # 取数据
    content1 = models.information.objects.order_by('-Time').all()[: 1000]
    content1_list = []
    for i in content1:
        if i.Content:
            content1_list.append(i.Content)
    # 构建标签列表，长度必须匹配也是300条
    with open('static/txt/labels.txt', 'r', encoding='utf-8') as f:
        labels = [line.strip() for line in f]
    df = pd.DataFrame({'content': content1_list, 'label': labels})
    # print(df)
    # jieba分词停用词，不要全禁了否则无法训练模型
    # 下载停用词列表（只需要下载一次）
    nltk.download('stopwords')
    # 加载停用词列表
    stop_words = set(stopwords.words('chinese'))

    # 定义分词函数，并去除停用词
    def cut_words(text):
        if text:
            words = jieba.cut(text)
            result = []
            for word in words:
                if word not in stop_words:
                    result.append(word)
            return ' '.join(result)

    # 对文本进行分词，并去除停用词
    df['content_cut'] = df['content'].apply(cut_words)
    # print(df)
    # 使用TF-IDF提取文本特征
    # 尝试设置 min_df 或者 max_df 参数
    # min_df 表示词汇的最小文档频率，即词汇在多少个文档中出现过；
    # max_df 表示词汇的最大文档频率，即词汇在多少个文档中过于常见。
    # 这两个参数的默认值分别为 1 和 1.0，如果设置得太小或者太大，
    # 都可能导致出现词汇表为空的情况
    # X 是一个稀疏矩阵，表示经过特征提取后的文本特征向量
    tfidf_vectorizer = TfidfVectorizer(min_df=1)
    X = tfidf_vectorizer.fit_transform(df['content_cut'])
    # 定义标签向量，并进行训练集和测试集的划分
    # 将标签向量定义为 y，并使用 train_test_split 方法将数据集划分为训练集和测试集
    y = df['label']
    # test_size 参数指定测试集所占的比例，random_state 参数指定随机数种子，以便结果可重复
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    # 定义分类器，并进行训练和预测
    # 使用 MultinomialNB 类创建一个朴素贝叶斯分类器，并使用 fit 方法对训练集进行训练，使用 predict 方法对测试集进行预测
    clf = MultinomialNB()
    clf.fit(X_train, y_train)
    y_pred = clf.predict(X_test)
    # 计算模型的准确率等评价指标。使用 scikit-learn 库提供的评价指标函数，可以计算模型的准确率、精确率、召回率、F1 值等指标
    # 评估分类器并计算指标
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred, average='weighted')
    recall = recall_score(y_test, y_pred, average='weighted')
    f1 = f1_score(y_test, y_pred, average='weighted')
    # print(accuracy, precision, recall, f1)
    # 计算混淆矩阵和分类报告
    cm = confusion_matrix(y_test, y_pred)
    cr = classification_report(y_test, y_pred)
    # 计算 ROC 曲线和 AUC 值
    fpr = dict()
    tpr = dict()
    roc_auc = dict()
    n_classes = len(set(y_test))
    y_test_bin = label_binarize(y_test, classes=list(set(y_test)))
    y_pred_bin = clf.predict_proba(X_test)
    for i in range(n_classes):
        fpr[i], tpr[i], _ = roc_curve(y_test_bin[:, i], y_pred_bin[:, i])
        roc_auc[i] = auc(fpr[i], tpr[i])
    # 绘制 ROC 曲线
    plt.figure()
    lw = 2
    colors = ['blue', 'red', 'green', 'orange']
    for i, color in zip(range(n_classes), colors):
        plt.plot(fpr[i], tpr[i], color=color, lw=lw, label='ROC curve of class {0} (AUC = {1:0.2f})'
                                                           ''.format(i, roc_auc[i]))
    plt.plot([0, 1], [0, 1], 'k--', lw=lw)
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver operating characteristic')
    plt.legend(loc="lower right")
    plt.savefig('static/img/roc_curve.png')
    plt.close()
    # 计算 MAP 和 NDCG 值
    y_test_bin = label_binarize(y_test, classes=list(set(y_test)))
    y_pred_bin = clf.predict_proba(X_test)
    ap = average_precision_score(y_test_bin, y_pred_bin, average='weighted')
    ndcg = ndcg_score(y_test_bin, y_pred_bin)
    # 将指标作为上下文变量传递给模板
    result1 = {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'cm': cm,
        'cr': cr,
        'auc': np.mean(list(roc_auc.values())),
        'map': ap,
        'ndcg': ndcg,
    }
    # print(cm)
    return render(request, 'display_2.html', {'result1': result1})


from sklearn.linear_model import LinearRegression
import datetime


# 信息可视化大屏跳转到“回归预测”
def display_3(request):
    # 获取历史数据
    content2 = models.information.objects.annotate(date=Substr("Time", 1, 7)).values("date").annotate(
        count=Count("id")).order_by("-date").all()
    content2_ganzhou = models.information.objects.filter(City="赣州市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("赣州市：", content2_ganzhou)
    content2_guigang = models.information.objects.filter(City="贵港市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("贵港市：", content2_guigang)
    content2_huizhou = models.information.objects.filter(City="惠州市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("惠州市：",content2_huizhou)
    content2_jining = models.information.objects.filter(City="济宁市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("济宁市：",content2_jining)
    content2_luzhou = models.information.objects.filter(City="泸州市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("泸州市：",content2_luzhou)
    content2_nanyang = models.information.objects.filter(City="南阳市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("南阳市：",content2_nanyang)
    content2_neijiang = models.information.objects.filter(City="内江市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("内江市：",content2_neijiang)
    content2_shiyan = models.information.objects.filter(City="十堰市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("十堰市：",content2_shiyan)
    content2_wenzhou = models.information.objects.filter(City="温州市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("温州市：",content2_wenzhou)
    content2_yongzhou = models.information.objects.filter(City="永州市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("永州市：",content2_yongzhou)
    content2_zhaoqing = models.information.objects.filter(City="肇庆市").annotate(date=Substr("Time", 1, 7)).values(
        "date").annotate(count=Count("id")).order_by("-date").all()
    # print("肇庆市：",content2_zhaoqing)
    # 按日期从过去到现在排序
    content2 = list(content2)
    content2 = [elem for elem in content2 if elem['date'] is not None]
    content2.sort(key=lambda x: x['date'])
    content2_ganzhou = list(content2_ganzhou)
    content2_ganzhou = [elem for elem in content2_ganzhou if elem['date'] is not None]
    content2_ganzhou.sort(key=lambda x: x['date'])
    content2_guigang = list(content2_guigang)
    content2_guigang = [elem for elem in content2_guigang if elem['date'] is not None]
    content2_guigang.sort(key=lambda x: x['date'])
    content2_huizhou = list(content2_huizhou)
    content2_huizhou = [elem for elem in content2_huizhou if elem['date'] is not None]
    content2_huizhou.sort(key=lambda x: x['date'])
    content2_jining = list(content2_jining)
    content2_jining = [elem for elem in content2_jining if elem['date'] is not None]
    content2_jining.sort(key=lambda x: x['date'])
    content2_luzhou = list(content2_luzhou)
    content2_luzhou = [elem for elem in content2_luzhou if elem['date'] is not None]
    content2_luzhou.sort(key=lambda x: x['date'])
    content2_nanyang = list(content2_nanyang)
    content2_nanyang = [elem for elem in content2_nanyang if elem['date'] is not None]
    content2_nanyang.sort(key=lambda x: x['date'])
    content2_neijiang = list(content2_neijiang)
    content2_neijiang = [elem for elem in content2_neijiang if elem['date'] is not None]
    content2_neijiang.sort(key=lambda x: x['date'])
    content2_shiyan = list(content2_shiyan)
    content2_shiyan = [elem for elem in content2_shiyan if elem['date'] is not None]
    content2_shiyan.sort(key=lambda x: x['date'])
    content2_wenzhou = list(content2_wenzhou)
    content2_wenzhou = [elem for elem in content2_wenzhou if elem['date'] is not None]
    content2_wenzhou.sort(key=lambda x: x['date'])
    content2_yongzhou = list(content2_yongzhou)
    content2_yongzhou = [elem for elem in content2_yongzhou if elem['date'] is not None]
    content2_yongzhou.sort(key=lambda x: x['date'])
    content2_zhaoqing = list(content2_zhaoqing)
    content2_zhaoqing = [elem for elem in content2_zhaoqing if elem['date'] is not None]
    content2_zhaoqing.sort(key=lambda x: x['date'])
    # 将日期字符串解析为 datetime 对象
    dates = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2]
    dates_ganzhou = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_ganzhou]
    dates_guigang = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_guigang]
    dates_huizhou = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_huizhou]
    dates_jining = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_jining]
    dates_luzhou = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_luzhou]
    dates_nanyang = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_nanyang]
    dates_neijiang = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_neijiang]
    dates_shiyan = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_shiyan]
    dates_wenzhou = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_wenzhou]
    dates_yongzhou = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_yongzhou]
    dates_zhaoqing = [datetime.datetime.strptime(item['date'], '%Y-%m') for item in content2_zhaoqing]
    # 将 datetime 对象转换为数字类型的时间戳，并转换为二维数组
    timestamps = np.array([(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates]).reshape(-1, 1)
    timestamps_ganzhou = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_ganzhou]).reshape(-1, 1)
    timestamps_guigang = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_guigang]).reshape(-1, 1)
    timestamps_huizhou = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_huizhou]).reshape(-1, 1)
    timestamps_jining = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_jining]).reshape(-1, 1)
    timestamps_luzhou = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_luzhou]).reshape(-1, 1)
    timestamps_nanyang = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_nanyang]).reshape(-1, 1)
    timestamps_neijiang = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_neijiang]).reshape(-1, 1)
    timestamps_shiyan = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_shiyan]).reshape(-1, 1)
    timestamps_wenzhou = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_wenzhou]).reshape(-1, 1)
    timestamps_yongzhou = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_yongzhou]).reshape(-1, 1)
    timestamps_zhaoqing = np.array(
        [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in dates_zhaoqing]).reshape(-1, 1)
    # 将日期转换为数值类型
    values = np.array([item['count'] for item in content2])
    values_ganzhou = np.array([item['count'] for item in content2_ganzhou])
    values_guigang = np.array([item['count'] for item in content2_guigang])
    values_huizhou = np.array([item['count'] for item in content2_huizhou])
    values_jining = np.array([item['count'] for item in content2_jining])
    values_luzhou = np.array([item['count'] for item in content2_luzhou])
    values_nanyang = np.array([item['count'] for item in content2_nanyang])
    values_neijiang = np.array([item['count'] for item in content2_neijiang])
    values_shiyan = np.array([item['count'] for item in content2_shiyan])
    values_wenzhou = np.array([item['count'] for item in content2_wenzhou])
    values_yongzhou = np.array([item['count'] for item in content2_yongzhou])
    values_zhaoqing = np.array([item['count'] for item in content2_zhaoqing])
    # 创建线性回归模型并拟合数据
    model = LinearRegression()
    model.fit(timestamps, values)
    model_ganzhou = LinearRegression()
    model_ganzhou.fit(timestamps_ganzhou, values_ganzhou)
    model_guigang = LinearRegression()
    model_guigang.fit(timestamps_guigang, values_guigang)
    model_huizhou = LinearRegression()
    model_huizhou.fit(timestamps_huizhou, values_huizhou)
    model_jining = LinearRegression()
    model_jining.fit(timestamps_jining, values_jining)
    model_luzhou = LinearRegression()
    model_luzhou.fit(timestamps_luzhou, values_luzhou)
    model_nanyang = LinearRegression()
    model_nanyang.fit(timestamps_nanyang, values_nanyang)
    model_neijiang = LinearRegression()
    model_neijiang.fit(timestamps_neijiang, values_neijiang)
    model_shiyan = LinearRegression()
    model_shiyan.fit(timestamps_shiyan, values_shiyan)
    model_wenzhou = LinearRegression()
    model_wenzhou.fit(timestamps_wenzhou, values_wenzhou)
    model_yongzhou = LinearRegression()
    model_yongzhou.fit(timestamps_yongzhou, values_yongzhou)
    model_zhaoqing = LinearRegression()
    model_zhaoqing.fit(timestamps_zhaoqing, values_zhaoqing)

    # 预测未来一年的值
    future_dates = [datetime.datetime(year=2022, month=month, day=1) for month in range(1, 13)]
    future_timestamps = [(date - datetime.datetime(1970, 1, 1)).total_seconds() for date in future_dates]
    predicted_values = model.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_ganzhou = model_ganzhou.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_guigang = model_guigang.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_huizhou = model_huizhou.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_jining = model_jining.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_luzhou = model_luzhou.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_nanyang = model_nanyang.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_neijiang = model_neijiang.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_shiyan = model_shiyan.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_wenzhou = model_wenzhou.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_yongzhou = model_yongzhou.predict(np.array(future_timestamps).reshape(-1, 1))
    predicted_values_zhaoqing = model_zhaoqing.predict(np.array(future_timestamps).reshape(-1, 1))
    # 保存预测未来一年结果
    predict = []
    for i, value in enumerate(predicted_values):
        date_str = future_dates[i].strftime('%Y-%m')
        predict.append({'date': date_str, 'value': int(value)})
    predict_ganzhou = []
    for i, value in enumerate(predicted_values_ganzhou):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_ganzhou.append({'date': date_str, 'value': int(value)})
    predict_guigang = []
    for i, value in enumerate(predicted_values_guigang):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_guigang.append({'date': date_str, 'value': int(value)})
    predict_huizhou = []
    for i, value in enumerate(predicted_values_huizhou):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_huizhou.append({'date': date_str, 'value': int(value)})
    predict_jining = []
    for i, value in enumerate(predicted_values_jining):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_jining.append({'date': date_str, 'value': int(value)})
    predict_luzhou = []
    for i, value in enumerate(predicted_values_luzhou):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_luzhou.append({'date': date_str, 'value': int(value)})
    predict_nanyang = []
    for i, value in enumerate(predicted_values_nanyang):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_nanyang.append({'date': date_str, 'value': int(value)})
    predict_neijiang = []
    for i, value in enumerate(predicted_values_neijiang):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_neijiang.append({'date': date_str, 'value': int(value)})
    predict_shiyan = []
    for i, value in enumerate(predicted_values_shiyan):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_shiyan.append({'date': date_str, 'value': int(value)})
    predict_wenzhou = []
    for i, value in enumerate(predicted_values_wenzhou):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_wenzhou.append({'date': date_str, 'value': int(value)})
    predict_yongzhou = []
    for i, value in enumerate(predicted_values_yongzhou):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_yongzhou.append({'date': date_str, 'value': int(value)})
    predict_zhaoqing = []
    for i, value in enumerate(predicted_values_zhaoqing):
        date_str = future_dates[i].strftime('%Y-%m')
        predict_zhaoqing.append({'date': date_str, 'value': int(value)})

    # 打印预测结果
    # print("all：",predict)
    # print("赣州市：", predict_ganzhou)
    # print("贵港市：", predict_guigang)
    # print("惠州市：", predict_huizhou)
    # print("济宁市：", predict_jining)
    # print("泸州市：", predict_luzhou)
    # print("南阳市：", predict_nanyang)
    # print("内江市：", predict_neijiang)
    # print("十堰市：", predict_shiyan)
    # print("温州市：", predict_wenzhou)
    # print("永州市：", predict_yongzhou)
    # print("肇庆市：", predict_zhaoqing)
    # 计算模型的 R2 分数
    r2_score = model.score(timestamps, values)
    r2_score_ganzhou = model_ganzhou.score(timestamps_ganzhou, values_ganzhou)
    r2_score_guigang = model_guigang.score(timestamps_guigang, values_guigang)
    r2_score_huizhou = model_huizhou.score(timestamps_huizhou, values_huizhou)
    r2_score_jining = model_jining.score(timestamps_jining, values_jining)
    r2_score_luzhou = model_luzhou.score(timestamps_luzhou, values_luzhou)
    r2_score_nanyang = model_nanyang.score(timestamps_nanyang, values_nanyang)
    r2_score_neijiang = model_neijiang.score(timestamps_neijiang, values_neijiang)
    r2_score_shiyan = model_shiyan.score(timestamps_shiyan, values_shiyan)
    r2_score_wenzhou = model_wenzhou.score(timestamps_wenzhou, values_wenzhou)
    r2_score_yongzhou = model_yongzhou.score(timestamps_yongzhou, values_yongzhou)
    r2_score_zhaoqing = model_zhaoqing.score(timestamps_zhaoqing, values_zhaoqing)
    # 获取模型的系数和截距
    coef = model.coef_
    intercept = model.intercept_
    coef_ganzhou = model_ganzhou.coef_
    intercept_ganzhou = model_ganzhou.intercept_
    coef_guigang = model_guigang.coef_
    intercept_guigang = model_guigang.intercept_
    coef_huizhou = model_huizhou.coef_
    intercept_huizhou = model_huizhou.intercept_
    coef_jining = model_jining.coef_
    intercept_jining = model_jining.intercept_
    coef_luzhou = model_luzhou.coef_
    intercept_luzhou = model_luzhou.intercept_
    coef_nanyang = model_nanyang.coef_
    intercept_nanyang = model_nanyang.intercept_
    coef_neijiang = model_neijiang.coef_
    intercept_neijiang = model_neijiang.intercept_
    coef_shiyan = model_shiyan.coef_
    intercept_shiyan = model_shiyan.intercept_
    coef_wenzhou = model_wenzhou.coef_
    intercept_wenzhou = model_wenzhou.intercept_
    coef_yongzhou = model_yongzhou.coef_
    intercept_yongzhou = model_yongzhou.intercept_
    coef_zhaoqing = model_zhaoqing.coef_
    intercept_zhaoqing = model_zhaoqing.intercept_
    # 将predict列表转化为json
    predict_json = json.dumps(predict)
    predict_json_ganzhou = json.dumps(predict_ganzhou)
    predict_json_guigang = json.dumps(predict_guigang)
    predict_json_huizhou = json.dumps(predict_huizhou)
    predict_json_jining = json.dumps(predict_jining)
    predict_json_luzhou = json.dumps(predict_luzhou)
    predict_json_nanyang = json.dumps(predict_nanyang)
    predict_json_neijiang = json.dumps(predict_neijiang)
    predict_json_shiyan = json.dumps(predict_shiyan)
    predict_json_wenzhou = json.dumps(predict_wenzhou)
    predict_json_yongzhou = json.dumps(predict_yongzhou)
    predict_json_zhaoqing = json.dumps(predict_zhaoqing)
    result2 = {
        'r2_score': r2_score,
        'coef': coef,
        'intercept': intercept
    }
    result2_ganzhou = {
        'r2_score_ganzhou': r2_score_ganzhou,
        'coef_ganzhou': coef_ganzhou,
        'intercept_ganzhou': intercept_ganzhou
    }
    result2_guigang = {
        'r2_score_guigang': r2_score_guigang,
        'coef_guigang': coef_guigang,
        'intercept_guigang': intercept_guigang
    }
    result2_huizhou = {
        'r2_score_huizhou': r2_score_huizhou,
        'coef_huizhou': coef_huizhou,
        'intercept_huizhou': intercept_huizhou
    }
    result2_jining = {
        'r2_score_jining': r2_score_jining,
        'coef_jining': coef_jining,
        'intercept_jining': intercept_jining
    }
    result2_luzhou = {
        'r2_score_luzhou': r2_score_luzhou,
        'coef_luzhou': coef_luzhou,
        'intercept_luzhou': intercept_luzhou
    }
    result2_nanyang = {
        'r2_score_nanyang': r2_score_nanyang,
        'coef_nanyang': coef_nanyang,
        'intercept_nanyang': intercept_nanyang
    }
    result2_neijiang = {
        'r2_score_neijiang': r2_score_neijiang,
        'coef_neijiang': coef_neijiang,
        'intercept_neijiang': intercept_neijiang
    }
    result2_shiyan = {
        'r2_score_shiyan': r2_score_shiyan,
        'coef_shiyan': coef_shiyan,
        'intercept_shiyan': intercept_shiyan
    }
    result2_wenzhou = {
        'r2_score_wenzhou': r2_score_wenzhou,
        'coef_wenzhou': coef_wenzhou,
        'intercept_wenzhou': intercept_wenzhou
    }
    result2_yongzhou = {
        'r2_score_yongzhou': r2_score_yongzhou,
        'coef_yongzhou': coef_yongzhou,
        'intercept_yongzhou': intercept_yongzhou
    }
    result2_zhaoqing = {
        'r2_score_zhaoqing': r2_score_zhaoqing,
        'coef_zhaoqing': coef_zhaoqing,
        'intercept_zhaoqing': intercept_zhaoqing
    }
    return render(
        request, 'display_3.html', {
            'predict_json': predict_json, 'result2': result2,
            'predict_json_ganzhou': predict_json_ganzhou, 'result2_ganzhou': result2_ganzhou,
            'predict_json_guigang': predict_json_guigang, 'result2_guigang': result2_guigang,
            'predict_json_huizhou': predict_json_huizhou, 'result2_huizhou': result2_huizhou,
            'predict_json_jining': predict_json_jining, 'result2_jining': result2_jining,
            'predict_json_luzhou': predict_json_luzhou, 'result2_luzhou': result2_luzhou,
            'predict_json_nanyang': predict_json_nanyang, 'result2_nanyang': result2_nanyang,
            'predict_json_neijiang': predict_json_neijiang, 'result2_neijiang': result2_neijiang,
            'predict_json_shiyan': predict_json_shiyan, 'result2_shiyan': result2_shiyan,
            'predict_json_wenzhou': predict_json_wenzhou, 'result2_wenzhou': result2_wenzhou,
            'predict_json_yongzhou': predict_json_yongzhou, 'result2_yongzhou': result2_yongzhou,
            'predict_json_zhaoqing': predict_json_zhaoqing, 'result2_zhaoqing': result2_zhaoqing,
        }
    )
